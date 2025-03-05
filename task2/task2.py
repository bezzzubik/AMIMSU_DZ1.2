from multiprocessing import Value
import time
import openpyxl as oxl
from openpyxl.chart import LineChart, Reference
import random
import copy

excel_file = 'NU.xlsx'
try:
    df = oxl.load_workbook(excel_file)
except FileNotFoundError:
    print(f"File not found")
    exit()

sheet = df.active
Start_mass=[]

for row in sheet.iter_rows(min_row=1, max_row = 15, values_only=True):
    s = list(row)
    for i in range(15):
        s[i] = float(s[i])
    Start_mass.append(s)

mass = copy.deepcopy(Start_mass)

mass.insert(0, [0, 1,2,3,4,5,6,7,8,9,10,11,12,13,14,15])

for i in range(1,16):
    mass[i].insert(0, i)

for row in mass:
    for el in row:
        print(f"{el:7.2f}", end ="")
    print()
print()
print()

Ig = 1


def progg():
    for i in range(1, 16):

        for j in range(i+1, 16):
            if mass[j][i] != 0:
                #Перемещаем до ближайшего ненулевого значения или до диагонального
                lj = j
                while i != lj and mass[lj-1][i] == 0:
                    for li in range(16):
                        a = mass[li][lj]
                        mass[li][lj] = mass[li][lj-1]
                        mass[li][lj-1] = a
                    for li in range(16):
                        a = mass[lj][li]
                        mass[lj][li] = mass[lj-1][li]
                        mass[lj-1][li] = a
                    lj -= 1


        for j in range(i+1, 16):
            if mass[i][j] != 0:
                #Перемещаем до ближайшего ненулевого значения или до диагонального
                lj = j
                while i != lj and mass[i][lj-1] == 0:
                    for li in range(16):
                        a = mass[li][lj]
                        mass[li][lj] = mass[li][lj-1]
                        mass[li][lj-1] = a
                    for li in range(16):
                        a = mass[lj][li]
                        mass[lj][li] = mass[lj-1][li]
                        mass[lj-1][li] = a
                    lj -= 1
    return


def print_mass():
    for row in mass:
        for el in row:
            print(f"{el:7.2f}", end ="")
        print()
    print()
    print()
    return


progg()
print(f"MASSIV: ")
print("-----------------------------------------------------------------------------")
print_mass()
print("-----------------------------------------------------------------------------")


ll = 0
index_max = 0
pMax = 0
s = 0
for elem in mass:
    p = 0
    k = 0
    for i in range(1, 16):
        if elem[i] == 0:
            if k == 1:
                p+=1
        else:
            k = 1
            if p > pMax:
                pMax = p
                index_max = s
            p = 0
    s += 1


i = 1
j = i
lj = index_max
while i != lj:
    for li in range(16):
        a = mass[li][lj]
        mass[li][lj] = mass[li][lj-1]
        mass[li][lj-1] = a
    for li in range(16):
        a = mass[lj][li]
        mass[lj][li] = mass[lj-1][li]
        mass[lj-1][li] = a
    lj -= 1

progg()
print_mass()


sheet = df.create_sheet("Obrabotka")
for row, itm in enumerate(mass, start=1):
    for column, item in enumerate(itm, start=1):
        sheet.cell(row = row, column = column).value = item



#Выделяем 3 класса существенных состояний, считаем Pt-E 
Mass1 = copy.deepcopy(mass)


#Иммитационка
Pp = []
Pp = [row[1:] for row in mass[1:]]

for i in range(15):
    prd_zn = 0
    for j in range(15):
        if Pp[i][j] != 0:
            Pp[i][j] += prd_zn
            prd_zn = Pp[i][j]

for row in Pp:
    for el in row:
        print(f"{el:7.2f}", end ="")
    print()
print()
print()



random.seed(time.time())
mass_count_trans = []
mass_trans = []

for IStart in range(15):
    for ll in range(10):
        i = IStart
        mass_trans_local = [i+1]
        mass_count_trans_local = [0 for i in range(15)]
        for lll in range(100):
            stp = random.random()
            j = 0
            while j < 14 and Pp[i][j] <= stp:
                j+=1
            i = j
            mass_count_trans_local[i] += 1
            mass_trans_local.append(i+1)

        mass_trans.append(mass_trans_local)
        mass_count_trans.append(mass_count_trans_local)

    sheet = df.create_sheet("count_trans " + str(IStart+1))
    for row, itm in enumerate(mass_count_trans, start=1):
        for column, item in enumerate(itm, start=1):
            sheet.cell(row = row, column = column).value = item

    sheet = df.create_sheet("Trans " + str(IStart+1))
    for row, itm in enumerate(mass_trans, start=1):
        for column, item in enumerate(itm, start=1):
            sheet.cell(row = row, column = column).value = item

    for col in range(1, 16):
        sheet.cell(row = 11, column = col).value = col
    mass_trans.clear()
    mass_count_trans.clear()


df.save("Itog.xlsx")